#!/usr/bin/env python3
"""
Lifestyles at Crystal Springs II — Budget Comparison Builder
Reads HOA budget PDFs, normalizes line items, and generates:
  1) Excel workbook: pivot-ready data + cross-tab pivot + change log + summaries
  2) One-page trustee accountability brief PDF
  3) Homeowner-friendly summary graphic PNG

Requirements:
  pip install pandas numpy openpyxl matplotlib reportlab pdfplumber pymupdf

Optional OCR (needed for image-based PDFs like scanned tables):
  pip install pytesseract pillow
  Install Tesseract OCR engine:
    - Windows: https://github.com/UB-Mannheim/tesseract/wiki
    - macOS: brew install tesseract
    - Linux: sudo apt-get install tesseract-ocr

Usage:
  python build_budget_deliverables.py --pdfs 2023.pdf 2024.pdf 2025.pdf 2026.pdf --outdir output
"""

import argparse
import os
import re
from pathlib import Path

import numpy as np
import pandas as pd

# Deliverables
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule

from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib import colors
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

import pdfplumber

# OCR fallback for image-based PDFs
OCR_AVAILABLE = False
try:
    import fitz  # PyMuPDF
    from PIL import Image
    import pytesseract
    OCR_AVAILABLE = True
except Exception:
    OCR_AVAILABLE = False


# ---------------------------
# Helpers: parsing amounts
# ---------------------------
def parse_money(x: str) -> float:
    if x is None:
        return 0.0
    s = str(x).strip()
    if s in ("", "-", "—"):
        return 0.0
    # handle negatives like (1,234.56)
    neg = False
    if s.startswith("(") and s.endswith(")"):
        neg = True
        s = s[1:-1]
    s = s.replace("$", "").replace(",", "").strip()
    try:
        v = float(s)
        return -v if neg else v
    except ValueError:
        return 0.0


def detect_year_from_filename(path: str) -> int | None:
    m = re.search(r"(20\d{2})", Path(path).name)
    return int(m.group(1)) if m else None


# ---------------------------
# Extraction strategies
# ---------------------------
def extract_text_pdfplumber(pdf_path: str) -> str:
    """Extract text from a PDF if it’s a real text PDF."""
    parts = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t.strip():
                parts.append(t)
    return "\n".join(parts).strip()


def extract_text_ocr(pdf_path: str, dpi: int = 200) -> str:
    """
    OCR fallback: render each page to an image and OCR it.
    Only runs if OCR_AVAILABLE is True.
    """
    if not OCR_AVAILABLE:
        raise RuntimeError("OCR requested but pytesseract/PyMuPDF not available.")

    doc = fitz.open(pdf_path)
    out = []
    for i, page in enumerate(doc):
        pix = page.get_pixmap(dpi=dpi)
        img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)
        text = pytesseract.image_to_string(img)
        out.append(text)
    return "\n".join(out).strip()


def extract_budget_text(pdf_path: str, try_ocr: bool = True) -> tuple[str, bool]:
    """
    Returns: (text, used_ocr)
    """
    t = extract_text_pdfplumber(pdf_path)
    if t and len(t) > 200:
        return t, False
    if try_ocr and OCR_AVAILABLE:
        t2 = extract_text_ocr(pdf_path)
        return t2, True
    return t, False


# ---------------------------
# Parsing the budgets
# ---------------------------
LINE_RE = re.compile(
    r"^\s*(?P<acct>\d{4,6}\s*-\s*)?(?P<item>.+?)\s+\$?(?P<a>\(?-?[\d,]+\.\d{2}\)?)\s+\$?(?P<b>\(?-?[\d,]+\.\d{2}\)?)\s+\$?(?P<c>\(?-?[\d,]+\.\d{2}\)?)\s+\$?(?P<d>\(?-?[\d,]+\.\d{2}\)?)\s*$"
)

def parse_budget_table(text: str, year: int) -> pd.DataFrame:
    """
    Heuristic parser that tries to pull line items from the common “Projected/Approved” style tables.
    Many HOA management budgets follow that structure.
    Output columns:
      Year, Type (Income/Expense), Category, Item, Amount
    Amount is the "Approved <year>" column when possible; otherwise best guess.

    If your PDFs have different layouts, this is the part to tweak.
    """

    # Normalize text
    lines = [l.strip() for l in text.splitlines() if l.strip()]
    rows = []

    current_type = None   # Income / Expense
    current_cat = None

    # Some budgets label categories explicitly; we also infer by account ranges / keywords.
    def set_type_from_line(line: str):
        nonlocal current_type
        if re.search(r"\bIncome\b", line, re.IGNORECASE):
            current_type = "Income"
        if re.search(r"\bExpense\b", line, re.IGNORECASE):
            current_type = "Expense"

    for line in lines:
        set_type_from_line(line)

        # Category headings often appear as plain text lines
        if re.match(r"^(Total\s+)?(General|Transition|Utilities|Landscaping|Contracts|Fund Contributions|Maintenance|Insurance)", line, re.IGNORECASE):
            # Don’t treat "Total ..." as a category heading
            if not line.lower().startswith("total"):
                current_cat = line.strip()
            continue

        # Try match main table row with 4 numeric columns
        m = LINE_RE.match(line)
        if m:
            item = m.group("item").strip()
            a = parse_money(m.group("a"))
            b = parse_money(m.group("b"))
            c = parse_money(m.group("c"))
            d = parse_money(m.group("d"))

            # Heuristic: "Approved <year>" typically the last numeric column (d)
            amount = d

            # Infer type if missing
            t = current_type
            if t is None:
                # crude guess: income items often contain 'Fees', 'Income', 'Reimbursement'
                if re.search(r"(Fees|Income|Reimbursement|Contribution)", item, re.IGNORECASE):
                    t = "Income"
                else:
                    t = "Expense"

            # Infer category if missing
            cat = current_cat
            if cat is None:
                if re.search(r"(Management|Audit|Legal|Postage|Office|Meeting|Annual)", item, re.IGNORECASE):
                    cat = "General & Administrative"
                elif re.search(r"Transition", item, re.IGNORECASE):
                    cat = "Transition"
                elif re.search(r"Electric", item, re.IGNORECASE):
                    cat = "Utilities"
                elif re.search(r"Landscape|Mulch|Irrigation", item, re.IGNORECASE):
                    cat = "Landscaping"
                elif re.search(r"Trash|Snow", item, re.IGNORECASE):
                    cat = "Contracts"
                elif re.search(r"Insurance", item, re.IGNORECASE):
                    cat = "Insurance"
                elif re.search(r"Reserve", item, re.IGNORECASE):
                    cat = "Reserves"
                else:
                    cat = "Other"

            # Skip totals in itemized list (we build totals ourselves)
            if item.lower().startswith("total "):
                continue

            rows.append({"Year": year, "Type": t, "Category": cat, "Item": item, "Amount": float(amount)})
            continue

    df = pd.DataFrame(rows)
    if df.empty:
        return df

    # Standardize category labels a bit
    df["Category"] = df["Category"].replace({
        "Total General & Administrative": "General & Administrative",
        "Total Utilities": "Utilities",
        "Total Landscaping": "Landscaping",
        "Total Contracts": "Contracts",
        "Total Fund Contributions": "Reserves",
        "Reserve Contribution": "Reserves",
    })

    return df


def normalize_items(df: pd.DataFrame) -> pd.DataFrame:
    """
    Normalize item names so similar lines match across years.
    """
    def norm(s: str) -> str:
        s = re.sub(r"\s+", " ", s.strip())
        s = s.replace("&", "and")
        # collapse some common variations
        s = re.sub(r"\bNJ\b", "NJ", s, flags=re.IGNORECASE)
        s = re.sub(r"Printing\s+and\s+Copying", "Printing & Copying", s, flags=re.IGNORECASE)
        s = re.sub(r"General\s+Maint\..*Repair", "General Maint. & Repair", s, flags=re.IGNORECASE)
        return s

    df = df.copy()
    df["Item"] = df["Item"].astype(str).map(norm)
    df["Category"] = df["Category"].astype(str).map(norm)
    return df


# ---------------------------
# Derived tables: pivots + change flags
# ---------------------------
def build_pivot_flags(df_long: pd.DataFrame, years: list[int]) -> pd.DataFrame:
    pt = (
        df_long
        .pivot_table(index=["Type", "Category", "Item"], columns="Year", values="Amount", aggfunc="sum", fill_value=0.0)
        .reindex(columns=years, fill_value=0.0)
        .reset_index()
    )

    # Added/removed by year (0 -> nonzero, nonzero -> 0)
    def first_nonzero(row):
        for y in years:
            if row.get(y, 0.0) != 0.0:
                return y
        return ""

    def last_nonzero(row):
        for y in reversed(years):
            if row.get(y, 0.0) != 0.0:
                return y
        return ""

    added = []
    removed = []
    for _, r in pt.iterrows():
        add_year = ""
        rem_year = ""
        prev = 0.0
        for y in years:
            cur = float(r[y])
            if prev == 0.0 and cur != 0.0:
                add_year = y if add_year == "" else add_year
            if prev != 0.0 and cur == 0.0:
                rem_year = y if rem_year == "" else rem_year
            prev = cur
        added.append(add_year)
        removed.append(rem_year)

    pt["AddedIn"] = added
    pt["RemovedIn"] = removed
    pt["FirstNonzero"] = pt.apply(first_nonzero, axis=1)
    pt["LastNonzero"] = pt.apply(last_nonzero, axis=1)
    return pt


# ---------------------------
# Excel writer (openpyxl)
# ---------------------------
def write_excel(df_data: pd.DataFrame, pivot_flags: pd.DataFrame, years: list[int], out_path: str):
    wb = Workbook()
    wb.remove(wb.active)

    header_font = Font(bold=True)
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def autosize(ws, max_width=60):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                if cell.value is None:
                    continue
                max_len = max(max_len, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_len + 2, max_width)

    def style_header(ws, freeze="A2"):
        ws.freeze_panes = freeze
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

    def apply_borders(ws):
        for r in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for c in r:
                c.border = border
                c.alignment = Alignment(vertical="top", wrap_text=True)

    # Data sheet
    ws = wb.create_sheet("Data")
    for r in dataframe_to_rows(df_data, index=False, header=True):
        ws.append(r)
    style_header(ws, "A2")
    amt_col = list(df_data.columns).index("Amount") + 1
    for col in ws.iter_cols(min_col=amt_col, max_col=amt_col, min_row=2):
        for c in col:
            c.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    apply_borders(ws)
    autosize(ws)

    # Make Excel Table
    ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
    tab = Table(displayName="BudgetData", ref=ref)
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)

    # Pivot sheet (prebuilt cross-tab)
    ws = wb.create_sheet("Pivot")
    pf = pivot_flags.copy()
    pf = pf[["Type", "Category", "Item"] + years + ["AddedIn", "RemovedIn", "FirstNonzero", "LastNonzero"]]
    for r in dataframe_to_rows(pf, index=False, header=True):
        ws.append(r)
    style_header(ws, "D2")
    apply_borders(ws)

    # Currency formatting for year columns
    year_start = 4
    for i in range(len(years)):
        col = year_start + i
        for cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cells:
                c.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    # Conditional formatting: Added/Removed
    added_col = 4 + len(years)
    removed_col = added_col + 1
    green_fill = PatternFill("solid", fgColor="D9EAD3")
    red_fill = PatternFill("solid", fgColor="F4CCCC")
    ws.conditional_formatting.add(
        f"{get_column_letter(added_col)}2:{get_column_letter(added_col)}{ws.max_row}",
        CellIsRule(operator="notEqual", formula=['""'], fill=green_fill),
    )
    ws.conditional_formatting.add(
        f"{get_column_letter(removed_col)}2:{get_column_letter(removed_col)}{ws.max_row}",
        CellIsRule(operator="notEqual", formula=['""'], fill=red_fill),
    )

    # Highlight 0->nonzero and nonzero->0 changes year-to-year
    light_green = PatternFill("solid", fgColor="E7F3E7")
    light_red = PatternFill("solid", fgColor="FBE4E4")
    for i in range(1, len(years)):
        prev_col = year_start + i - 1
        cur_col = year_start + i
        prev_letter = get_column_letter(prev_col)
        cur_letter = get_column_letter(cur_col)
        rng = f"{cur_letter}2:{cur_letter}{ws.max_row}"
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f"AND({prev_letter}2=0,{cur_letter}2<>0)"], fill=light_green))
        ws.conditional_formatting.add(rng, FormulaRule(formula=[f"AND({prev_letter}2<>0,{cur_letter}2=0)"], fill=light_red))

    autosize(ws)

    # Change Log
    ws = wb.create_sheet("Change Log")
    changes = pivot_flags[(pivot_flags["AddedIn"] != "") | (pivot_flags["RemovedIn"] != "")].copy()
    changes = changes[["Type", "Category", "Item"] + years + ["AddedIn", "RemovedIn"]]
    for r in dataframe_to_rows(changes, index=False, header=True):
        ws.append(r)
    style_header(ws, "D2")
    apply_borders(ws)
    for i in range(len(years)):
        col = 4 + i
        for cells in ws.iter_cols(min_col=col, max_col=col, min_row=2):
            for c in cells:
                c.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    autosize(ws)

    # ReadMe
    ws = wb.create_sheet("ReadMe")
    ws["A1"] = "How to use this workbook"
    ws["A1"].font = Font(bold=True, size=14)
    lines = [
        "1) Data sheet is a clean table (BudgetData) you can pivot from in Excel.",
        "2) Excel: Insert → PivotTable → select BudgetData → New Worksheet.",
        "   Suggested pivot:",
        "   - Rows: Category, Item",
        "   - Columns: Year",
        "   - Values: Sum of Amount",
        "   - Filter: Type (Income vs Expense)",
        "3) Pivot sheet is a prebuilt cross-tab with AddedIn/RemovedIn flags and change highlighting."
    ]
    for i, line in enumerate(lines, start=3):
        ws[f"A{i}"] = line
    ws.column_dimensions["A"].width = 120

    wb.save(out_path)


# ---------------------------
# Homeowner graphic
# ---------------------------
def write_homeowner_graphic(df_long: pd.DataFrame, years: list[int], out_path: str):
    exp = df_long[df_long["Type"] == "Expense"].copy()
    # Standard rollup
    cat_totals = exp.pivot_table(index="Year", columns="Category", values="Amount", aggfunc="sum", fill_value=0.0)
    cat_totals = cat_totals.reindex(index=years, fill_value=0.0)

    # Choose a stable category order (adjust as desired)
    cat_order = [c for c in [
        "General & Administrative", "Transition", "Landscaping", "Contracts",
        "Insurance", "Utilities", "General Maintenance", "Reserves", "Other"
    ] if c in cat_totals.columns]

    x = np.arange(len(years))
    bottom = np.zeros(len(years))

    plt.figure(figsize=(12, 7))
    for cat in cat_order:
        vals = cat_totals[cat].values
        plt.bar(x, vals, bottom=bottom, label=cat)
        bottom += vals

    totals_exp = bottom
    plt.xticks(x, [str(y) for y in years])
    plt.ylabel("Annual Operating Expenses ($)")
    plt.title("Where HOA Operating Money Goes — by Category")
    plt.legend(ncol=2, fontsize=9)

    for i, total in enumerate(totals_exp):
        plt.text(i, total, f"${total:,.0f}", ha="center", va="bottom", fontsize=9)

    plt.tight_layout()
    plt.savefig(out_path, dpi=200)
    plt.close()


# ---------------------------
# Trustee brief (one-page PDF)
# ---------------------------
def write_trustee_brief(df_long: pd.DataFrame, years: list[int], out_path: str):
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("title", parent=styles["Title"], fontSize=16, spaceAfter=10)
    h_style = ParagraphStyle("h", parent=styles["Heading2"], fontSize=11, spaceBefore=6, spaceAfter=4)
    b_style = ParagraphStyle("b", parent=styles["BodyText"], fontSize=9, leading=11)
    small_style = ParagraphStyle("small", parent=styles["BodyText"], fontSize=8, leading=10, textColor=colors.grey)

    doc = SimpleDocTemplate(
        out_path,
        pagesize=letter,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
        topMargin=0.6 * inch, bottomMargin=0.5 * inch
    )

    story = []
    story.append(Paragraph(f"Trustee Accountability Brief — Lifestyles at Crystal Springs II ({years[0]}–{years[-1]} Budgets)", title_style))
    story.append(Paragraph(
        "Purpose: a one-page checklist of big budget shifts, recurring cost drivers, and document requests homeowners can make.",
        b_style
    ))

    # Biggest YOY change in latest year
    exp = df_long[df_long["Type"] == "Expense"].copy()
    latest = years[-1]
    prev = years[-2] if len(years) >= 2 else None

    if prev is not None:
        pt = exp.pivot_table(index=["Category", "Item"], columns="Year", values="Amount", aggfunc="sum", fill_value=0.0)
        if prev in pt.columns and latest in pt.columns:
            yoy = (pt[latest] - pt[prev]).sort_values(ascending=False)
            top_changes = yoy.head(8)

            story.append(Spacer(1, 6))
            story.append(Paragraph(f"Big swings in the latest budget ({prev} → {latest})", h_style))
            bullets = "<br/>".join([
                f"• {idx[0]} / {idx[1]}: ${v:,.0f} ({'increase' if v > 0 else 'decrease'})"
                for idx, v in top_changes.items()
                if abs(v) >= 1
            ])
            story.append(Paragraph(bullets or "• No significant changes detected.", b_style))

    story.append(Spacer(1, 6))
    story.append(Paragraph("Recurring / structural items to explain", h_style))
    rec = [
        "• “Transition” costs appearing across multiple years: ask what is still “transitioning,” the end date, and yearly deliverables/invoices.",
        "• Reserve contribution level: ask whether it matches a current reserve study and a 30-year funding plan.",
        "• Largest cost drivers: landscaping/contracts — ask for contract scope changes and bid history."
    ]
    story.append(Paragraph("<br/>".join(rec), b_style))

    story.append(Spacer(1, 6))
    story.append(Paragraph("Document requests homeowners can make (in writing)", h_style))
    qs = [
        "• Vendor backup: contracts + last 3 bids for Landscaping, Snow, Trash, Insurance.",
        "• Legal spend: engagement letters + description of matters covered by legal fees.",
        "• Reserve compliance: latest reserve study date, preparer, and funding plan.",
        "• Meeting minutes: where budget was presented, discussed, and adopted."
    ]
    story.append(Paragraph("<br/>".join(qs), b_style))

    story.append(Spacer(1, 6))
    story.append(Paragraph("Compliance watchlist (high-level; not legal advice)", h_style))
    comp = [
        "• Budget notice/process: confirm owners received the proposed budget and meeting notice per governing docs and applicable NJ requirements.",
        "• Reserves: confirm a current reserve study / funding plan exists if required for your association type.",
        "• Records access: confirm owners can inspect financial records and minutes per governing documents/statute."
    ]
    story.append(Paragraph("<br/>".join(comp), b_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph("Tip: Ask for a written timeline + responsible party if answers are delayed.", small_style))

    doc.build(story)


# ---------------------------
# Main
# ---------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--pdfs", nargs="+", required=True, help="List of budget PDFs (2023..2026 etc)")
    ap.add_argument("--outdir", default="output", help="Output directory")
    ap.add_argument("--no-ocr", action="store_true", help="Disable OCR fallback")
    args = ap.parse_args()

    outdir = Path(args.outdir)
    outdir.mkdir(parents=True, exist_ok=True)

    all_dfs = []
    sources = []

    for pdf in args.pdfs:
        year = detect_year_from_filename(pdf)
        if year is None:
            raise ValueError(f"Could not infer year from filename: {pdf}")

        text, used_ocr = extract_budget_text(pdf, try_ocr=(not args.no_ocr))
        if not text or len(text) < 200:
            print(f"[WARN] Low/empty extracted text for {pdf}. used_ocr={used_ocr}. You may need OCR.")
        df = parse_budget_table(text, year)

        if df.empty:
            print(f"[WARN] Parsed zero rows from {pdf}. You likely need OCR or a custom table parser for that layout.")
        else:
            df["SourcePDF"] = Path(pdf).name
            df["UsedOCR"] = used_ocr
            all_dfs.append(df)

    if not all_dfs:
        raise RuntimeError("No data parsed from PDFs. Enable OCR or adjust parser.")

    df_long = pd.concat(all_dfs, ignore_index=True)
    df_long = normalize_items(df_long)

    years = sorted(df_long["Year"].unique().tolist())

    # Build pivot flags table
    pivot_flags = build_pivot_flags(df_long, years)

    # Write deliverables
    excel_path = str(outdir / f"Budget_Comparison_{years[0]}-{years[-1]}.xlsx")
    png_path = str(outdir / f"Homeowner_Summary_Graphic_{years[0]}-{years[-1]}.png")
    pdf_path = str(outdir / f"Trustee_Accountability_Brief_{years[0]}-{years[-1]}.pdf")

    # Order columns for Excel Data table
    df_data = df_long[["Year", "Type", "Category", "Item", "Amount", "SourcePDF", "UsedOCR"]].sort_values(
        ["Type", "Category", "Item", "Year"]
    )

    write_excel(df_data, pivot_flags, years, excel_path)
    write_homeowner_graphic(df_long, years, png_path)
    write_trustee_brief(df_long, years, pdf_path)

    print("Wrote:")
    print(" -", excel_path)
    print(" -", png_path)
    print(" -", pdf_path)


if __name__ == "__main__":
    main()
